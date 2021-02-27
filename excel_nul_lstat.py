# -*- coding: utf-8 -*-

# ---------------------前提条件----------------
# 1、excel第1行为表头，数据从第2行开始。
# 2、数据表为excel文件的第一个sheet。
# ---------------------------------------------
# 脚本用途：提取Excel文件中每列（字段）中数据个数，从而统计每个字段的信息缺失率，以表中有效行数为基准;
# 支持直接输入文件夹路径，文件夹下所有excel文件的统计，也可统计单个excel表的数据
# 传参规则：文件夹路径，示例：python excel_stat.py 文件夹绝对路径/文件绝对路径
# 输出文件：传参路径文件夹下生成stat_result.xlsx文件。
# 输出格式：文件名，文件总行数，文件表头（列字段），各字段统计数据
# 输出示例：
# | test    | 数据总行数为：27 | ... |
# | 表中字段  | 工号 | 姓名 | 性别 |
# | 字段空值数|  22  |  6  |  5  |
# | 字段空值率| 81%  | 22% | 19% |

import xlrd
import xlsxwriter
import openpyxl
import sys
import time
import os

def read_excel(path):
    # print(path)
    wb = xlrd.open_workbook(path,encoding_override='utf-8')  # 打开原始文件
    # 默认excel文件的第一个sheet为数据表
    sheet = wb.sheet_by_index(0)
    global nrows
    # 默认excel第1行为表头，数据从第2行开始
    nrows = sheet.nrows - 1
    global biaotou_list
    biaotou_list = sheet.row_values(0)
    global col_num_list
    global col_rat_list
    for i in range(0,len(biaotou_list)):
        col_value=sheet.col_values(i)
        # print(col_value)
        j=0
        for x in col_value:
            # print(x)
            if not x:
                j=j+1
        # print(j)
        # 检查工号/姓名字段的数据是否完整
        if j!=0 and col_value[0] == "工号":
            print("工号字段的行数与表的最大行数不同，请检查")
        if j!=0 and col_value[0] == "姓名":
            print("姓名字段的行数与表的最大行数不同，请检查")
        percent=int((j)/nrows*100)
        # print(percent)
        col_num_list.append(j)
        col_rat_list.append(str(percent)+"%")

def write_excel(file_path,nrows,biaotou_list,col_num_list,col_rat_list):
    filename = os.path.basename(file_path).split('.')[-2]
    dir_path = os.path.dirname(file_path)+r'\stat_result.xlsx'
    # 如果excel文件已存在则更新或新建sheet，如果文件不存在则新建excel文件
    if os.path.exists(dir_path):
        wb = openpyxl.load_workbook(dir_path)
        # 更新已存在的sheet：删除已存在的sheet，重新添加sheet
        if filename in wb:
            # 获取已存在sheet所在位置，保证新建时顺序不乱
            pos = wb.sheetnames.index(filename)
            del wb[filename]
            ws = wb.create_sheet(title=filename, index=pos)
        else:
            ws = wb.create_sheet(title=filename)
        ws['A1']= filename
        ws['B1'] = '数据总行数为：'+str(nrows)
        ws['A2'] = "表中字段"
        for i in range(1,len(biaotou_list)+1):
            ws.cell(row=2,column=i+1).value=biaotou_list[i-1]
        ws['A3'] = "字段空值数"
        for i in range(1,len(col_num_list)+1):
            ws.cell(row=3,column=i+1).value=col_num_list[i-1]
        ws['A4'] = "字段空值率"
        for i in range(1,len(col_rat_list)+1):
            ws.cell(row=4,column=i+1).value=col_rat_list[i-1]
        if 'stat_result' in wb:
            del wb['stat_result']
        wb.save(dir_path)
    else:
        workbook = xlsxwriter.Workbook(dir_path)
        addsheet = workbook.add_worksheet(filename)
        addsheet.write('A1', filename)
        addsheet.write('B1', '数据总行数为：'+str(nrows))
        addsheet.write('A2', '表中字段')
        addsheet.write_row('B2', biaotou_list)
        addsheet.write('A3', '字段空值数')
        addsheet.write_row('B3', col_num_list)
        addsheet.write('A4', '字段空值率')
        addsheet.write_row('B4', col_rat_list)
        workbook.close()

if __name__ == '__main__':
    start = time.perf_counter()
    path=sys.argv[1]
    nrows = 0
    biaotou_list = []
    col_num_list = []
    col_rat_list = []
    # 如果路径参数为目录，则集中处理文件夹内各个excel表数据
    if os.path.isdir(path):
        for filename in os.listdir(path):
            if filename.split('.')[-1] == 'xlsx' or filename.split('.')[-1] == 'xls':
                file_path = os.path.join(path, filename)
                nrows = 0
                biaotou_list = []
                col_num_list = []
                col_rat_list = []
                read_excel(file_path)
                write_excel(file_path, nrows, biaotou_list, col_num_list, col_rat_list)
                print('总行数为' + str(nrows))
                print('各字段缺失情况：')
                print(biaotou_list)
                print(col_num_list)
                print(col_rat_list)
    else:
        read_excel(path)
        write_excel(path, nrows, biaotou_list, col_num_list, col_rat_list)
        print('总行数为' + str(nrows))
        print('各字段缺失情况：')
        print(biaotou_list)
        print(col_num_list)
        print(col_rat_list)
    end = time.perf_counter()
    print('Running time: %s Seconds' % (end - start))