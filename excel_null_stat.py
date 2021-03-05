# -*- coding: utf-8 -*-

# ---------------------前提条件----------------
# 1、excel第1行为表头，数据从第2行开始。
# 2、数据表为excel文件的第一个sheet。
# ---------------------------------------------
# 脚本用途：提取Excel文件中每列（字段）中数据个数，从而统计每个字段的信息缺失率，以表中有效行数为基准;
# 支持直接输入文件夹路径，文件夹下所有excel文件的统计，也可统计单个excel表的数据
# 传参规则：文件夹路径，示例：python excel_stat.py 文件夹绝对路径/文件绝对路径
# 输出文件：传参路径文件夹下生成stat_result.xlsx文件。
# 输出格式：文件名，文件总行数，文件表头（列字段），各字段统计数据
# 输出示例：
# | test    | 数据总行数为：27 | ... |
# | 表中字段  | 工号 | 姓名 | 性别 |
# | 字段空值数|  22  |  6  |  5  |
# | 字段空值率| 81%  | 22% | 19% |

import xlrd
import xlwt
import xlsxwriter
import openpyxl
from openpyxl.styles import Border, Side, colors
from xlutils.copy import copy
import sys
import time
import os

def read_excel_xlrd(path):
    # print(path)
    wb = xlrd.open_workbook(path,encoding_override='utf-8')  # 打开原始文件
    # 默认excel文件的第一个sheet为数据表
    sheet = wb.sheet_by_index(0)
    global nrows
    # 默认excel第1行为表头，数据从第2行开始
    nrows = sheet.nrows - 1
    global biaotou_list
    biaotou_list = sheet.row_values(0)
    global col_num_list
    global col_rat_list
    for i in range(0,len(biaotou_list)):
        col_value=sheet.col_values(i)
        # print(col_value)
        j=0
        for x in col_value:
            # print(x)
            if (not x) or x.isspace():
                j=j+1
        # print(j)
        # 检查工号/姓名字段的数据是否完整
        if j!=0 and col_value[0] == "工号":
            print("工号字段的行数与表的最大行数不同，请检查")
        if j!=0 and col_value[0] == "姓名":
            print("姓名字段的行数与表的最大行数不同，请检查")
        percent=int(j/nrows*100)
        # print(percent)
        col_num_list.append(j)
        col_rat_list.append(str(percent)+"%")

def read_excel_openpyxl(path):
    # print(path)
    wb = openpyxl.load_workbook(path)  # 打开原始文件
    # active默认指向excel文件的第一个sheet
    sheet = wb.active
    global nrows
    # 默认excel第1行为表头，数据从第2行开始
    nrows = sheet.max_row - 1
    global biaotou_list
    # 通常sheet.iter_rows返回为cell对象(tuple类型)组成的generator生成器类型(values_only默认为False)
    # 当指定values_only=True，返回为cell.value(tuple类型)组成的generator生成器类型
    # 由于参数指定为第一行，所以此循环实际只循环一次，但是由于sheet.iter_rows()返回为生成器类型，所以只能使用for循环取出实际的tuple，再转化为list类型
    for row_value_tuple in sheet.iter_rows(1,1,values_only=True):
        biaotou_list = list(row_value_tuple)
    global col_num_list
    global col_rat_list
    # 从第一列到第len(biaotou_list)列，从第二行开始，因为第一行为表头。
    for col_value_tuple in sheet.iter_cols(1,len(biaotou_list),2,values_only=True):
        j=0
        for x in col_value_tuple:
            # print(x)
            if (not x) or str(x).isspace():
                j=j+1
        # print(j)
        # 检查工号/姓名字段的数据是否完整
        if j!=0 and col_value_tuple[0] == "工号":
            print("工号字段的行数与表的最大行数不同，请检查")
        if j!=0 and col_value_tuple[0] == "姓名":
            print("姓名字段的行数与表的最大行数不同，请检查")
        percent=int(j/nrows*100)
        # print(percent)
        col_num_list.append(j)
        col_rat_list.append(str(percent)+"%")

# 简化版：每个输入excel分别生成一个输出excel,不更新已有文件，相同目录下生成“原文件名+result.xlsx”文件
def write_excel_simple(file_path,nrows,biaotou_list,col_rat_list):
    filename = os.path.basename(file_path).split('.')[-2]
    result_name = filename+'result.xlsx'
    dir_path = os.path.dirname(file_path)+'\\'+result_name
    workbook = xlsxwriter.Workbook(dir_path)
    addsheet = workbook.add_worksheet(filename)
    head_str = filename + "（记录条数：" + str(nrows) + "，提交人：            ，接收人：          ）"
    addsheet.merge_range('A1:N1',head_str)
    head_str = filename + "（记录条数：" + str(nrows) + "，提交人：            ，接收人：          ）"
    addsheet.write('A1', head_str)
    addsheet.write('A2', '字段名称')
    addsheet.write_row('B2', biaotou_list)
    addsheet.write('A3', '空值率')
    addsheet.write_row('B3', col_rat_list)
    addsheet.write('A4', '数据说明：')
    workbook.close()

# 更新+sheet版：将当前文件夹只存在一个stat_result.xlsx文件，每个sheet对应本文件夹的每个输入excel；stat_result.xlsx文件可更新
def write_excel(file_path,nrows,biaotou_list,col_rat_list):
    filename = os.path.basename(file_path).split('.')[-2]
    dir_path = os.path.dirname(file_path)+r'\stat_result.xlsx'
    # 如果excel文件已存在则更新或新建sheet，如果文件不存在则新建excel文件
    if os.path.exists(dir_path):
        wb = openpyxl.load_workbook(dir_path)
        # 更新已存在的sheet：删除已存在的sheet，重新添加sheet
        if filename in wb:
            # 获取已存在sheet所在位置，保证新建时顺序不乱
            pos = wb.sheetnames.index(filename)
            del wb[filename]
            ws = wb.create_sheet(title=filename, index=pos)
        else:
            ws = wb.create_sheet(title=filename)
        ws.merge_cells('A1:N1')
        head_str = filename + "（记录条数："+ str(nrows) + "，提交人：            ，接收人：          ）"
        ws['A1']= head_str
        ws['A2'] = "字段名称"
        for i in range(1,len(biaotou_list)+1):
            ws.cell(row=2,column=i+1).value=biaotou_list[i-1]
        ws['A3'] = "空值率"
        for i in range(1,len(col_rat_list)+1):
            ws.cell(row=3,column=i+1).value=col_rat_list[i-1]
        ws['A4'] = "数据说明："
        # 更新时有可能将上次已生成的stat_result当做数据，处理成统计文件的sheet，所以要删除此多余的sheet
        if 'stat_result' in wb:
            del wb['stat_result']
        wb.save(dir_path)
    else:
        workbook = xlsxwriter.Workbook(dir_path)
        addsheet = workbook.add_worksheet(filename)
        head_str = filename + "（记录条数：" + str(nrows) + "，提交人：            ，接收人：          ）"
        addsheet.merge_range('A1:N1', head_str)
        addsheet.write('A1', head_str)
        addsheet.write('A2', '字段名称')
        addsheet.write_row('B2', biaotou_list)
        addsheet.write('A3', '空值率')
        addsheet.write_row('B3', col_rat_list)
        addsheet.write('A4', '数据说明：')
        workbook.close()

# 更新+sheet+格式版（openpyxl+xlsxwriter版）：在write_excel()基础上增加了输出格式的调整
def write_excel_format(file_path,nrows,biaotou_list,col_rat_list):
    filename = os.path.basename(file_path).split('.')[-2]
    dir_path = os.path.dirname(file_path)+r'\stat_result.xlsx'
    # 如果excel文件已存在则更新或新建sheet，如果文件不存在则新建excel文件
    if os.path.exists(dir_path):
        wb = openpyxl.load_workbook(dir_path)
        # 更新已存在的sheet：删除已存在的sheet，重新添加sheet
        if filename in wb:
            # 获取已存在sheet所在位置，保证新建时顺序不乱
            pos = wb.sheetnames.index(filename)
            del wb[filename]
            ws = wb.create_sheet(title=filename, index=pos)
        else:
            ws = wb.create_sheet(title=filename)
        ws.merge_cells('A1:N1')
        head_str = filename + "（记录条数："+ str(nrows) + "，提交人：            ，接收人：          ）"
        # 设置首行字体加粗
        ws['A1']= head_str
        font_set = openpyxl.styles.Font(name='宋体', size=12, bold=True)
        ws['A1'].font=font_set
        # 为统计数据行设置边框
        border_set = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin', color=colors.BLACK),
                            right=openpyxl.styles.Side(style='thin', color=colors.BLACK),
                            top=openpyxl.styles.Side(style='thin', color=colors.BLACK),
                            bottom=openpyxl.styles.Side(style='thin', color=colors.BLACK))
        ws['A2'] = "字段名称"
        ws['A2'].border = border_set
        for i in range(1,len(biaotou_list)+1):
            ws.cell(row=2,column=i+1).value=biaotou_list[i-1]
            ws.cell(row=2, column=i + 1).border = border_set
        ws['A3'] = "空值率"
        ws['A3'].border = border_set
        for i in range(1,len(col_rat_list)+1):
            ws.cell(row=3,column=i+1).value=col_rat_list[i-1]
            ws.cell(row=3, column=i + 1).border = border_set
        ws['A4'] = "数据说明："
        if 'stat_result' in wb:
            del wb['stat_result']
        # 设置合适的列宽：openpyxl的行或列的编号是从1开始的,openpyxl会根据字符数判断列宽，一个汉字大约等于2个字符,为防止文字紧贴边框，最后加1.5字符宽度
        for i in range(2,len(biaotou_list)+2):
            ws.column_dimensions[openpyxl.utils.cell.get_column_letter(i)].width = 2*len(biaotou_list[i-2])+1.5
        wb.save(dir_path)
    else:
        workbook = xlsxwriter.Workbook(dir_path)
        addsheet = workbook.add_worksheet(filename)
        # 合并单元格并设置首行字体加粗
        head_str = filename + "（记录条数：" + str(nrows) + "，提交人：            ，接收人：          ）"
        merge_format = workbook.add_format({'bold': True})
        addsheet.merge_range('A1:N1', head_str,merge_format)
        # 为统计数据行设置边框
        border_format = workbook.add_format({'border': 1})
        addsheet.write('A2', '字段名称', border_format)
        addsheet.write_row('B2', biaotou_list, border_format)
        addsheet.write('A3', '空值率', border_format)
        addsheet.write_row('B3', col_rat_list, border_format)
        addsheet.write('A4', '数据说明：')
        ############################################################################
        # 使用set_row()设置整行格式，会将无数据的部单元格也设置成目标格式
        # addsheet.set_row(2, None, border_format)
        ############################################################################
        # conditional_format为'type':'cell'后，缺少criteria总是报错
        # 根据表头长度计算出数据范围并转化为excel的列表示
        # x = xlsxwriter.utility.xl_col_to_name(len(biaotou_list))
        # pos = 'A2:'+ x +'3'
        # addsheet.conditional_format(pos, {'type':'cell','format': border_format})
        ############################################################################
        # 设置合适的列宽
        for i in range(1, len(biaotou_list) + 1):
            # 人工测试：11号宋体合适的列宽=1.875*汉字数+0.88
            addsheet.set_column(i, i, 1.875 * len(biaotou_list[i - 1]) + 0.88)
        workbook.close()

# 更新+sheet+格式版（xlrd+xlutils+xlwt版）：在write_excel()基础上增加了输出格式的调整
def write_excel_xlutils_xlwt(file_path,nrows,biaotou_list,col_rat_list):
    filename = os.path.basename(file_path).split('.')[-2]
    # 因为xlwt保存时只支持xls格式，所以只能更新xls文件，否则只能保存为一个同名的.xls文件，就相当于另存为了一份文件。
    dir_path = os.path.dirname(file_path)+r'\stat_result.xls'
    # 如果excel文件已存在则更新或新建sheet，如果文件不存在则新建excel文件
    if os.path.exists(dir_path):
        rd = xlrd.open_workbook(dir_path, formatting_info=True) # formatting_info=True，打开保留原有格式，这样保存时不会丢掉原有格式。
        # 使用xlutils.copy.copy()将xlrd.Book对象转化为xlwt.Workbook对象
        wb=copy(rd)
        # 更新已存在的sheet
        if filename in rd.sheet_names():
            # 获取已存在sheet
            ws = wb.get_sheet(filename)
        else:
            ws = wb.add_sheet(filename,cell_overwrite_ok=True) # cell_overwrite_ok=True保证同一单元格可以多次写入
    else:
        wb = xlwt.Workbook()
        ws = wb.add_sheet(filename, cell_overwrite_ok=True)
    head_str = filename + "（记录条数：" + str(nrows) + "，提交人：            ，接收人：          ）"
    style1=xlwt.XFStyle()
    font = xlwt.Font()
    font.name = '宋体'
    font.bold = True
    style1.font=font
    ws.write_merge(0,0,0,len(biaotou_list)-1,head_str,style1)
    style2 = xlwt.XFStyle()
    borders = xlwt.Borders()
    borders.left = xlwt.Borders.THIN
    borders.right = xlwt.Borders.THIN
    borders.top = xlwt.Borders.THIN
    borders.bottom = xlwt.Borders.THIN
    style2.borders = borders
    ws.write(1, 0, '字段名称：',style2)
    for i in range(0,len(biaotou_list)):
        ws.write(1, i+1, biaotou_list[i], style2)
    ws.write(2, 0, "空值率",style2)
    for i in range(0,len(col_rat_list)):
        ws.write(2, i+1, col_rat_list[i], style2)
    ws.write(3, 0, "数据说明：")
    # 调整列宽
    for i in range(0, len(biaotou_list)):
        # 256*字符数得到excel列宽,为了不显得特别紧凑添加两个字符宽度
        ws.col(i+1).width = 256*len(biaotou_list[i].encode('utf-8'))+2
    wb.save(dir_path)

# 更新+sheet+格式版（openpyxl版）：在write_excel()基础上增加了输出格式的调整
def write_excel_openpyxl(file_path,nrows,biaotou_list,col_rat_list):
    filename = os.path.basename(file_path).split('.')[-2]
    dir_path = os.path.dirname(file_path)+r'\stat_result.xlsx'
    # 如果excel文件已存在则更新或新建sheet，如果文件不存在则新建excel文件
    if os.path.exists(dir_path):
        wb = openpyxl.load_workbook(dir_path)
        # 更新已存在的sheet：删除已存在的sheet，重新添加sheet
        if filename in wb:
            # 获取已存在sheet所在位置，保证新建时顺序不乱
            pos = wb.sheetnames.index(filename)
            del wb[filename]
            ws = wb.create_sheet(title=filename, index=pos)
        else:
            ws = wb.create_sheet(title=filename)
    else:
        wb = openpyxl.Workbook()
        ws=wb.active
        ws.title=filename
    ws.merge_cells('A1:N1')
    head_str = filename + "（记录条数："+ str(nrows) + "，提交人：            ，接收人：          ）"
    # 设置首行字体加粗
    ws['A1']= head_str
    font_set = openpyxl.styles.Font(name='宋体', size=12, bold=True)
    ws['A1'].font=font_set
    # 为统计数据行设置边框
    border_set = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin', color=colors.BLACK),
                        right=openpyxl.styles.Side(style='thin', color=colors.BLACK),
                        top=openpyxl.styles.Side(style='thin', color=colors.BLACK),
                        bottom=openpyxl.styles.Side(style='thin', color=colors.BLACK))
    ws['A2'] = "字段名称"
    ws['A2'].border = border_set
    for i in range(1,len(biaotou_list)+1):
        ws.cell(row=2,column=i+1).value=biaotou_list[i-1]
        ws.cell(row=2, column=i + 1).border = border_set
    ws['A3'] = "空值率"
    ws['A3'].border = border_set
    for i in range(1,len(col_rat_list)+1):
        ws.cell(row=3,column=i+1).value=col_rat_list[i-1]
        ws.cell(row=3, column=i + 1).border = border_set
    ws['A4'] = "数据说明："
    # openpyxl的行或列的编号是从1开始的,openpyxl会根据字符数判断列宽，一个汉字大约等于2个字符,为防止文字紧贴边框，最后加1.5字符宽度
    for i in range(2,len(biaotou_list)+2):
        ws.column_dimensions[openpyxl.utils.cell.get_column_letter(i)].width = 2*len(biaotou_list[i-2])+1.5
    # sheet.set_width('col1', 30)
    wb.save(dir_path)



if __name__ == '__main__':
    start = time.perf_counter()
    path=sys.argv[1]
    nrows = 0
    biaotou_list = []
    col_num_list = []
    col_rat_list = []
    # 如果路径参数为目录，则集中处理文件夹内各个excel表数据
    if os.path.isdir(path):
        for filename in os.listdir(path):
            # 只读取excel文件，且需要将统计结果文件stat_result排除
            if filename.split('.')[-2]!="stat_result" and filename.split('.')[-1] == 'xlsx' or filename.split('.')[-1] == 'xls':
                file_path = os.path.join(path, filename)
                nrows = 0
                biaotou_list = []
                col_num_list = []
                col_rat_list = []
                # read_excel_xlrd(file_path)
                read_excel_openpyxl(file_path)
                # write_excel_simple(file_path, nrows, biaotou_list, col_rat_list)
                # write_excel(file_path, nrows, biaotou_list, col_rat_list)
                # write_excel_format(file_path, nrows, biaotou_list, col_rat_list)
                # write_excel_xlutils_xlwt(file_path, nrows, biaotou_list, col_rat_list)
                write_excel_openpyxl(file_path, nrows, biaotou_list, col_rat_list)
                print(filename+'总行数为' + str(nrows))
                print('各字段缺失情况：')
                print(biaotou_list)
                print(col_num_list)
                print(col_rat_list)
    else:
        # read_excel_xlrd(path)
        read_excel_openpyxl(path)
        # write_excel_simple(path, nrows, biaotou_list, col_rat_list)
        # write_excel(path, nrows, biaotou_list, col_rat_list)
        # write_excel_format(path, nrows, biaotou_list, col_rat_list)
        # write_excel_xlutils_xlwt(path, nrows, biaotou_list, col_rat_list)
        write_excel_openpyxl(path, nrows, biaotou_list, col_rat_list)
        print(path+'总行数为' + str(nrows))
        print('各字段缺失情况：')
        print(biaotou_list)
        print(col_num_list)
        print(col_rat_list)
    end = time.perf_counter()
    print('Running time: %s Seconds' % (end - start))